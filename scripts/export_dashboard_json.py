from __future__ import annotations

import argparse
import json
import math
import re
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR   = Path(__file__).resolve().parents[1]
INPUT_DIR  = BASE_DIR / "data" / "input"
OUTPUT_DIR = BASE_DIR / "dashboard" / "data"          # ← JSON lands here

ALL_TOKENS_PATH       = INPUT_DIR / "TAT - ALL TOKENS.xlsx"
COMPLETED_TOKENS_PATH = INPUT_DIR / "TAT - ALL COMPLETED TOKENS.xlsx"

DATA_SHEET = "DATA Daily Token Service Report"

# ---------------------------------------------------------------------------
# Lookup tables
# ---------------------------------------------------------------------------
MODALITY_LABELS = {
    "XR": "XRAY",
    "MR": "MRI",
    "CT": "CT",
    "US": "Ultrasound",
}

COMPLETED_STATUSES = {"Complete", "E-Complete"}

ALL_TOKEN_ALIASES = {
    "date":     ["Date"],
    "token":    ["Token"],
    "modality": ["Modality"],
    "status":   ["Status"],
}

COMPLETED_TOKEN_ALIASES = {
    "date":           ["Date"],
    "token":          ["Token"],
    "modality":       ["Modality"],
    "status":         ["Status"],
    "actual_tat":     ["ACTUAL Turnaround Time"],
    "target_tat":     ["TARGET TAT"],
    "dispatch_tat":   ["TAT - DISPATCH SERVICE"],
    "us_service_tat": ["TAT - US SERVICE", "U/S TAT"],
    "xr_service_tat": ["XRAY - TAT SERVICE", "Wait Time + Service time"],
    "ct_service_tat": ["TAT - CT SERVICE", "CT TAT"],
    "mr_service_tat": ["TAT - MRI SEVICE", "TAT - MRI SERVICE", "MRI TAT"],
    "us_billing_tat": ["TAT - U/S BILLING", "US BILLING TAT"],
    "xr_billing_tat": ["TAT - BILLING XRAY", "XRAY BILLING TAT"],
    "ct_billing_tat": ["TAT - CT BILLING", "CT BILLING TAT"],
    "mr_billing_tat": ["TAT - MRI BILLING", "MRI BILLING TAT"],
    "us_stage_wait":     [".ULTRASOUND - Wait Time"],
    "us_stage_service":  [".ULTRASOUND - Service Time"],
    "us_stage_hold":     [".ULTRASOUND - Hold Time"],
    "xr_stage_wait":     ["X-RAY - Wait Time"],
    "xr_stage_service":  ["X-RAY - Service Time"],
    "xr_stage_hold":     ["X-RAY - Hold Time"],
    "ct_stage_wait":     [".CT SCAN - Wait Time"],
    "ct_stage_service":  [".CT SCAN - Service Time"],
    "ct_stage_hold":     [".CT SCAN - Hold Time"],
    "mr_stage_wait":     [".MRI - Wait Time"],
    "mr_stage_service":  [".MRI - Service Time"],
    "mr_stage_hold":     [".MRI - Hold Time"],
    "us_billing_wait":    ["ULTRASOUND SERVICE - Wait Time"],
    "us_billing_service": ["ULTRASOUND SERVICE - Service Time"],
    "us_billing_hold":    ["ULTRASOUND SERVICE - Hold Time"],
    "xr_billing_wait":    ["X-RAY SERVICE - Wait Time"],
    "xr_billing_service": ["X-RAY SERVICE - Service Time"],
    "xr_billing_hold":    ["X-RAY SERVICE - Hold Time"],
    "ct_billing_wait":    ["CT SCAN SERVICE - Wait Time"],
    "ct_billing_service": ["CT SCAN SERVICE - Service Time"],
    "ct_billing_hold":    ["CT SCAN SERVICE - Hold Time"],
    "mr_billing_wait":    ["MRI SERVICE - Wait Time"],
    "mr_billing_service": ["MRI SERVICE - Service Time"],
    "mr_billing_hold":    ["MRI SERVICE - Hold Time"],
}

SERVICE_COLUMNS_BY_MODALITY = {
    "US": "us_service_tat",
    "XR": "xr_service_tat",
    "CT": "ct_service_tat",
    "MR": "mr_service_tat",
}

BILLING_COLUMNS_BY_MODALITY = {
    "US": "us_billing_tat",
    "XR": "xr_billing_tat",
    "CT": "ct_billing_tat",
    "MR": "mr_billing_tat",
}

# Modality code → display label (inverse of MODALITY_LABELS)
MODALITY_CODE_TO_LABEL = {v: k for k, v in MODALITY_LABELS.items()}

# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def clean_column_name(value: str) -> str:
    cleaned = re.sub(r"[^0-9a-zA-Z]+", "_", str(value).strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned.lower()


def normalize_status(value: object) -> str:
    if value is None or pd.isna(value):
        return "Unknown"
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return "Unknown"
    text = re.sub(r"\s*-\s*done$",     "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*-\s*not\s*done$", "", text, flags=re.IGNORECASE)
    text = text.replace("E. Complete", "E-Complete")
    text = text.replace("E Complete",  "E-Complete")
    text = text.replace("No Show",  "Noshow")
    text = text.replace("No-show",  "Noshow")
    text = text.replace("NosHow",   "Noshow")
    parts = [part.capitalize() for part in re.split(r"([\-/])", text) if part != ""]
    normalized = "".join(parts)
    return {
        "Noshow":     "Noshow",
        "E-complete": "E-Complete",
        "E-Complete": "E-Complete",
    }.get(normalized, normalized)


def normalize_modality(value: object) -> str:
    if value is None or pd.isna(value):
        return "Unknown"
    code = str(value).strip().upper()
    if not code or code == "NAN":
        return "Unknown"
    return MODALITY_LABELS.get(code, code)


def modality_code_from_label(label: str) -> str:
    return MODALITY_CODE_TO_LABEL.get(label, label)


def to_minutes(value: object) -> float:
    if value is None:
        return math.nan
    if isinstance(value, float) and math.isnan(value):
        return math.nan
    if isinstance(value, (int, float)):
        number = float(value)
        if 0 < number < 2:
            return number * 24 * 60
        return number
    if isinstance(value, pd.Timedelta):
        return value.total_seconds() / 60
    if isinstance(value, timedelta):
        return value.total_seconds() / 60
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute + value.second / 60
    if isinstance(value, time):
        return value.hour * 60 + value.minute + value.second / 60
    text = str(value).strip()
    if text in {"", "--", "nan", "NaN", "None", "0:00:00"}:
        return math.nan
    try:
        delta = pd.to_timedelta(text)
        return delta.total_seconds() / 60
    except Exception:
        pass
    try:
        number = float(text)
        if 0 < number < 2:
            return number * 24 * 60
        return number
    except ValueError:
        return math.nan


def minutes_to_hhmm(value: float) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "00:00"
    total_minutes = max(int(round(float(value))), 0)
    hours, minutes = divmod(total_minutes, 60)
    return f"{hours:02d}:{minutes:02d}"


def safe_number(value: float, digits: int = 2) -> float | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return round(float(value), digits)


def month_key(d: date) -> str:
    """Return a short lowercase month key, e.g. 'jan', 'feb', …"""
    return d.strftime("%b").lower()


def month_label(d: date) -> str:
    """Return 'January 2026' style label."""
    return d.strftime("%B %Y")


# ---------------------------------------------------------------------------
# Excel loading
# ---------------------------------------------------------------------------

def load_selected_columns(
    workbook_path: Path,
    sheet_name: str,
    column_aliases: dict[str, list[str]],
) -> pd.DataFrame:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        header_cells = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None
        )
        if not header_cells:
            return pd.DataFrame(columns=sorted(column_aliases))

        normalized_headers = {
            normalize_text(header): index
            for index, header in enumerate(header_cells)
            if header is not None and str(header).strip()
        }

        selected_columns: dict[str, int] = {}
        for output_name, aliases in column_aliases.items():
            for alias in aliases:
                normalized_alias = normalize_text(alias)
                if normalized_alias in normalized_headers:
                    selected_columns[output_name] = normalized_headers[normalized_alias]
                    break

        rows = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            record = {}
            for output_name, column_index in selected_columns.items():
                record[output_name] = row[column_index] if column_index < len(row) else None
            rows.append(record)

        frame = pd.DataFrame(rows)
        for output_name in column_aliases:
            if output_name not in frame.columns:
                frame[output_name] = pd.NA

        frame = frame[[column for column in column_aliases]]
        frame.columns = [clean_column_name(column) for column in frame.columns]
        return frame
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# Data preparation
# ---------------------------------------------------------------------------

def prepare_all_tokens(path: Path) -> pd.DataFrame:
    frame = load_selected_columns(path, DATA_SHEET, ALL_TOKEN_ALIASES)
    frame["date"]     = pd.to_datetime(frame["date"], errors="coerce").dt.date
    frame["token"]    = frame["token"].fillna("").astype(str).str.strip()
    frame["modality"] = frame["modality"].apply(normalize_modality)
    frame["status"]   = frame["status"].apply(normalize_status)
    return frame


def prepare_completed_tokens(path: Path) -> pd.DataFrame:
    frame = load_selected_columns(path, DATA_SHEET, COMPLETED_TOKEN_ALIASES)
    frame["date"]     = pd.to_datetime(frame["date"], errors="coerce").dt.date
    frame["token"]    = frame["token"].fillna("").astype(str).str.strip()
    frame["modality"] = frame["modality"].apply(normalize_modality)
    frame["status"]   = frame["status"].apply(normalize_status)

    frame["actual_tat_minutes"]   = frame["actual_tat"].apply(to_minutes)
    frame["target_tat_minutes"]   = frame["target_tat"].apply(to_minutes)
    frame["dispatch_tat_minutes"] = frame["dispatch_tat"].apply(to_minutes)

    for modality, column in SERVICE_COLUMNS_BY_MODALITY.items():
        frame[f"{column}_minutes"] = frame[column].apply(to_minutes)
    for modality, column in BILLING_COLUMNS_BY_MODALITY.items():
        frame[f"{column}_minutes"] = frame[column].apply(to_minutes)

    component_columns = [
        "us_stage_wait", "us_stage_service", "us_stage_hold",
        "xr_stage_wait", "xr_stage_service", "xr_stage_hold",
        "ct_stage_wait", "ct_stage_service", "ct_stage_hold",
        "mr_stage_wait", "mr_stage_service", "mr_stage_hold",
        "us_billing_wait", "us_billing_service", "us_billing_hold",
        "xr_billing_wait", "xr_billing_service", "xr_billing_hold",
        "ct_billing_wait", "ct_billing_service", "ct_billing_hold",
        "mr_billing_wait", "mr_billing_service", "mr_billing_hold",
    ]
    for column in component_columns:
        frame[f"{column}_minutes"] = frame[column].apply(to_minutes)

    return frame


# ---------------------------------------------------------------------------
# Builders — each returns a dict that becomes one JSON file
# ---------------------------------------------------------------------------

def _months_present(frame: pd.DataFrame) -> list[date]:
    """Return one representative date per calendar month, sorted ascending."""
    dates = frame["date"].dropna().unique()
    seen: dict[tuple[int, int], date] = {}
    for d in sorted(dates):
        key = (d.year, d.month)
        if key not in seen:
            seen[key] = d
    return list(seen.values())


def build_scorecards(all_tokens: pd.DataFrame) -> dict:
    """
    Per-month scorecard summary.
    Shape: { months: [{key, label, ...counts, completionRate}] }
    """
    months = _months_present(all_tokens)
    result = []
    for rep in months:
        key   = month_key(rep)
        label = month_label(rep)
        subset = all_tokens[
            (all_tokens["date"].apply(
                lambda d: d is not None and d.year == rep.year and d.month == rep.month
            ))
        ]
        status_counts   = subset["status"].value_counts()
        completed_total = int(status_counts.get("Complete",   0))
        ecomplete_total = int(status_counts.get("E-Complete", 0))
        total           = int(len(subset))
        result.append({
            "key":             key,
            "label":           label,
            "totalTokens":     total,
            "completedTokens": completed_total,
            "eCompleteTokens": ecomplete_total,
            "pendingTokens":   int(status_counts.get("Pending", 0)),
            "servingTokens":   int(status_counts.get("Serving", 0)),
            "noShowTokens":    int(status_counts.get("Noshow",  0)),
            "standbyTokens":   int(status_counts.get("Standby", 0)),
            "completionRate":  safe_number(
                (completed_total / total * 100) if total else 0.0, 1
            ),
        })
    return {"months": result}


def build_modality_status(all_tokens: pd.DataFrame) -> dict:
    """
    Per-month, per-modality completed vs non-completed counts.
    Shape: { months: [{key, label, modalities: [{modality, completed, nonCompleted, statusBreakdown}]}] }
    """
    months  = _months_present(all_tokens)
    cleaned = all_tokens[all_tokens["modality"] != "Unknown"].copy()
    result  = []
    for rep in months:
        subset = cleaned[
            cleaned["date"].apply(
                lambda d: d is not None and d.year == rep.year and d.month == rep.month
            )
        ]
        grouped = (
            subset.groupby(["modality", "status"], dropna=False)
            .size()
            .unstack(fill_value=0)
            .sort_index()
        )
        modalities = []
        for modality, row in grouped.iterrows():
            completed_count     = int(row.get("Complete", 0))
            non_completed_count = int(row.sum() - completed_count)
            status_breakdown    = {
                clean_column_name(s): int(c)
                for s, c in row.items() if int(c) > 0
            }
            modalities.append({
                "modality":       modality,
                "completed":      completed_count,
                "nonCompleted":   non_completed_count,
                "statusBreakdown": status_breakdown,
            })
        result.append({"key": month_key(rep), "label": month_label(rep), "modalities": modalities})
    return {"months": result}


def build_tat_vs_target(completed_tokens: pd.DataFrame) -> dict:
    """
    Per-month average actual vs target TAT by modality.
    Shape: { months: [{key, label, modalities: [{modality, actualMinutes, targetMinutes, ...}]}] }
    """
    months   = _months_present(completed_tokens)
    filtered = completed_tokens[
        completed_tokens["status"].isin(COMPLETED_STATUSES) &
        (completed_tokens["modality"] != "Unknown")
    ].copy()
    result = []
    for rep in months:
        subset = filtered[
            filtered["date"].apply(
                lambda d: d is not None and d.year == rep.year and d.month == rep.month
            )
        ]
        grouped = (
            subset.groupby("modality", dropna=False)
            .agg(
                actual_minutes=("actual_tat_minutes", "mean"),
                target_minutes=("target_tat_minutes", "mean"),
                token_count=("token", "count"),
            )
            .reset_index()
            .sort_values("modality")
        )
        modalities = []
        for row in grouped.itertuples(index=False):
            modalities.append({
                "modality":     row.modality,
                "actualMinutes": safe_number(row.actual_minutes),
                "targetMinutes": safe_number(row.target_minutes),
                "actualHHMM":   minutes_to_hhmm(row.actual_minutes),
                "targetHHMM":   minutes_to_hhmm(row.target_minutes),
                "tokenCount":   int(row.token_count),
            })
        result.append({"key": month_key(rep), "label": month_label(rep), "modalities": modalities})
    return {"months": result}


def build_tat_distribution(completed_tokens: pd.DataFrame) -> dict:
    """
    Per-month TAT split across billing / service / dispatch by modality.
    Shape: { months: [{key, label, modalities: [{modality, billingMinutes, ...}]}] }
    """
    months   = _months_present(completed_tokens)
    filtered = completed_tokens[
        completed_tokens["status"].isin(COMPLETED_STATUSES) &
        (completed_tokens["modality"] != "Unknown")
    ].copy()
    result = []
    for rep in months:
        subset = filtered[
            filtered["date"].apply(
                lambda d: d is not None and d.year == rep.year and d.month == rep.month
            )
        ]
        modalities = []
        for modality in sorted(subset["modality"].dropna().unique()):
            modality_code   = modality_code_from_label(modality)
            service_column  = SERVICE_COLUMNS_BY_MODALITY.get(modality_code)
            billing_column  = BILLING_COLUMNS_BY_MODALITY.get(modality_code)
            if not service_column or not billing_column:
                continue
            mf              = subset[subset["modality"] == modality]
            billing_mean    = mf[f"{billing_column}_minutes"].mean()
            service_mean    = mf[f"{service_column}_minutes"].mean()
            dispatch_mean   = mf["dispatch_tat_minutes"].mean()
            total_mean      = sum(v for v in [billing_mean, service_mean, dispatch_mean] if pd.notna(v))
            billing_pct  = (billing_mean  / total_mean * 100) if pd.notna(billing_mean)  and total_mean else math.nan
            service_pct  = (service_mean  / total_mean * 100) if pd.notna(service_mean)  and total_mean else math.nan
            dispatch_pct = (dispatch_mean / total_mean * 100) if pd.notna(dispatch_mean) and total_mean else math.nan
            modalities.append({
                "modality":       modality,
                "billingMinutes": safe_number(billing_mean),
                "billingHHMM":    minutes_to_hhmm(billing_mean),
                "billingPct":     safe_number(billing_pct,  1),
                "serviceMinutes": safe_number(service_mean),
                "serviceHHMM":    minutes_to_hhmm(service_mean),
                "servicePct":     safe_number(service_pct,  1),
                "dispatchMinutes": safe_number(dispatch_mean),
                "dispatchHHMM":   minutes_to_hhmm(dispatch_mean),
                "dispatchPct":    safe_number(dispatch_pct, 1),
                "tokenCount":     int(len(mf)),
            })
        result.append({"key": month_key(rep), "label": month_label(rep), "modalities": modalities})
    return {"months": result}


def build_daily_trends(completed_tokens: pd.DataFrame) -> dict:
    """
    Per-month daily actual vs target TAT trend by modality.
    Matches the shape of window.monthDailyTrendData from daily_trend_months.js.
    Shape: { months: [{key, label, modalities: {XR:[{date,completed,actualMin,...}], ...}}] }
    """
    months   = _months_present(completed_tokens)
    filtered = completed_tokens[
        completed_tokens["status"].isin(COMPLETED_STATUSES) &
        completed_tokens["date"].notna() &
        (completed_tokens["modality"] != "Unknown")
    ].copy()

    # Map display label → JS modality key used in the original file
    label_to_key = {v: k for k, v in MODALITY_LABELS.items()}

    result = []
    for rep in months:
        subset = filtered[
            filtered["date"].apply(
                lambda d: d is not None and d.year == rep.year and d.month == rep.month
            )
        ]
        grouped = (
            subset.groupby(["modality", "date"], dropna=False)
            .agg(
                tokens=("token", "count"),
                actual_minutes=("actual_tat_minutes", "mean"),
                target_minutes=("target_tat_minutes", "mean"),
            )
            .reset_index()
            .sort_values(["modality", "date"])
        )
        modalities_dict: dict[str, list] = {}
        for modality, mf in grouped.groupby("modality", sort=True):
            js_key = label_to_key.get(modality, modality)
            points = []
            for row in mf.itertuples(index=False):
                points.append({
                    "date":       row.date.isoformat() if isinstance(row.date, date) else str(row.date),
                    "completed":  int(row.tokens),
                    "actualMin":  safe_number(row.actual_minutes),
                    "actualHHMM": minutes_to_hhmm(row.actual_minutes),
                    "targetMin":  safe_number(row.target_minutes),
                    "targetHHMM": minutes_to_hhmm(row.target_minutes),
                })
            modalities_dict[js_key] = points
        result.append({
            "key":        month_key(rep),
            "label":      month_label(rep),
            "modalities": modalities_dict,
        })
    return {"months": result}


def build_daily_status_summary(all_tokens: pd.DataFrame) -> dict:
    """
    Per-month daily token counts by status + modality breakdown.
    Shape: { months: [{key, label, daily:[...], dailyModality:{...}}] }
    """
    months  = _months_present(all_tokens)
    cleaned = all_tokens.dropna(subset=["date"]).copy()
    result  = []
    for rep in months:
        subset = cleaned[
            cleaned["date"].apply(
                lambda d: d.year == rep.year and d.month == rep.month
            )
        ]
        grouped = (
            subset.groupby(["date", "status"], dropna=False)
            .size()
            .reset_index(name="count")
            .sort_values(["date", "status"])
        )
        daily_counts = []
        for date_val, date_frame in grouped.groupby("date"):
            date_str     = date_val.isoformat() if isinstance(date_val, date) else str(date_val)
            status_counts: dict[str, int] = {}
            total_tokens = 0
            for row in date_frame.itertuples(index=False):
                status = row.status if row.status else "Unknown"
                count  = int(row.count)
                status_counts[status] = count
                total_tokens         += count
            completed       = status_counts.get("Complete", 0)
            completion_rate = (completed / total_tokens * 100) if total_tokens > 0 else 0.0
            daily_counts.append({
                "date":           date_str,
                "totalTokens":    total_tokens,
                "statusCounts":   status_counts,
                "completedTokens": status_counts.get("Complete",   0),
                "eCompleteTokens": status_counts.get("E-Complete", 0),
                "pendingTokens":   status_counts.get("Pending",    0),
                "servingTokens":   status_counts.get("Serving",    0),
                "noShowTokens":    status_counts.get("Noshow",     0),
                "standbyTokens":   status_counts.get("Standby",    0),
                "completionRate":  safe_number(completion_rate, 1),
            })

        modality_grouped = (
            subset[subset["modality"] != "Unknown"]
            .groupby(["date", "modality", "status"], dropna=False)
            .size()
            .reset_index(name="count")
        )
        daily_modality: dict[str, dict] = {}
        for date_val, dmf in modality_grouped.groupby("date"):
            date_str = date_val.isoformat() if isinstance(date_val, date) else str(date_val)
            daily_modality[date_str] = {}
            for modality, mf in dmf.groupby("modality"):
                completed_count     = int(mf[mf["status"] == "Complete"]["count"].sum())
                non_completed_count = int(mf[mf["status"] != "Complete"]["count"].sum())
                daily_modality[date_str][modality] = {
                    "completed":    completed_count,
                    "nonCompleted": non_completed_count,
                }

        result.append({
            "key":          month_key(rep),
            "label":        month_label(rep),
            "daily":        daily_counts,
            "dailyModality": daily_modality,
        })
    return {"months": result}


def build_daily_process_breakdown(completed_tokens: pd.DataFrame) -> dict:
    """
    Per-month daily averages for service and billing stage components (wait/service/hold).
    Shape: { months: [{key, label, modalities: [{modality, points:[...]}]}] }
    """
    months   = _months_present(completed_tokens)
    filtered = completed_tokens[
        completed_tokens["status"].isin(COMPLETED_STATUSES) &
        completed_tokens["date"].notna() &
        (completed_tokens["modality"] != "Unknown")
    ].copy()

    stage_columns = {
        "Ultrasound": ("us_stage_wait_minutes",  "us_stage_service_minutes",  "us_stage_hold_minutes"),
        "XRAY":       ("xr_stage_wait_minutes",  "xr_stage_service_minutes",  "xr_stage_hold_minutes"),
        "CT":         ("ct_stage_wait_minutes",  "ct_stage_service_minutes",  "ct_stage_hold_minutes"),
        "MRI":        ("mr_stage_wait_minutes",  "mr_stage_service_minutes",  "mr_stage_hold_minutes"),
    }
    billing_columns = {
        "Ultrasound": ("us_billing_wait_minutes", "us_billing_service_minutes", "us_billing_hold_minutes"),
        "XRAY":       ("xr_billing_wait_minutes", "xr_billing_service_minutes", "xr_billing_hold_minutes"),
        "CT":         ("ct_billing_wait_minutes", "ct_billing_service_minutes", "ct_billing_hold_minutes"),
        "MRI":        ("mr_billing_wait_minutes", "mr_billing_service_minutes", "mr_billing_hold_minutes"),
    }

    result = []
    for rep in months:
        subset = filtered[
            filtered["date"].apply(
                lambda d: d is not None and d.year == rep.year and d.month == rep.month
            )
        ]
        modalities = []
        for modality in sorted(subset["modality"].dropna().unique()):
            modality_code      = modality_code_from_label(modality)
            service_column     = SERVICE_COLUMNS_BY_MODALITY.get(modality_code)
            billing_tat_column = BILLING_COLUMNS_BY_MODALITY.get(modality_code)
            if not service_column or not billing_tat_column:
                continue
            stage_cols   = stage_columns.get(modality)
            billing_cols = billing_columns.get(modality)
            mf = subset[subset["modality"] == modality].copy()
            grouped = (
                mf.groupby("date", dropna=False)
                .agg(
                    tokens=("token", "count"),
                    billing_minutes=(f"{billing_tat_column}_minutes", "mean"),
                    service_minutes=(f"{service_column}_minutes",     "mean"),
                    dispatch_minutes=("dispatch_tat_minutes",          "mean"),
                    service_wait=   (stage_cols[0],   "mean"),
                    service_service=(stage_cols[1],   "mean"),
                    service_hold=   (stage_cols[2],   "mean"),
                    billing_wait=   (billing_cols[0], "mean"),
                    billing_service=(billing_cols[1], "mean"),
                    billing_hold=   (billing_cols[2], "mean"),
                )
                .reset_index()
                .sort_values("date")
            )
            points = []
            for row in grouped.itertuples(index=False):
                points.append({
                    "date":           row.date.isoformat() if isinstance(row.date, date) else str(row.date),
                    "tokens":         int(row.tokens),
                    "billingMinutes": safe_number(row.billing_minutes),
                    "serviceMinutes": safe_number(row.service_minutes),
                    "dispatchMinutes": safe_number(row.dispatch_minutes),
                    "serviceBreakdown": {
                        "waitMinutes":    safe_number(row.service_wait),
                        "serviceMinutes": safe_number(row.service_service),
                        "holdMinutes":    safe_number(row.service_hold),
                    },
                    "billingBreakdown": {
                        "waitMinutes":    safe_number(row.billing_wait),
                        "serviceMinutes": safe_number(row.billing_service),
                        "holdMinutes":    safe_number(row.billing_hold),
                    },
                })
            modalities.append({"modality": modality, "points": points})
        result.append({"key": month_key(rep), "label": month_label(rep), "modalities": modalities})
    return {"months": result}


# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------

def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)
    print(f"  ✓  {path.relative_to(BASE_DIR)}")


def parse_iso_date(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(
            f"Invalid date '{value}'. Use YYYY-MM-DD format."
        ) from exc


def apply_date_filter(
    frame: pd.DataFrame,
    start_date: date | None,
    end_date: date | None,
) -> pd.DataFrame:
    filtered = frame
    if start_date is not None:
        filtered = filtered[filtered["date"] >= start_date]
    if end_date is not None:
        filtered = filtered[filtered["date"] <= end_date]
    return filtered


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export dashboard JSON files from TAT Excel sources."
    )
    parser.add_argument("--all-tokens",       dest="all_tokens_path",       default=str(ALL_TOKENS_PATH))
    parser.add_argument("--completed-tokens",  dest="completed_tokens_path",  default=str(COMPLETED_TOKENS_PATH))
    parser.add_argument("--output-dir",        dest="output_dir",             default=str(OUTPUT_DIR))
    parser.add_argument("--sheet-name",        dest="sheet_name",             default=DATA_SHEET)
    parser.add_argument("--start-date",        dest="start_date",             type=parse_iso_date, default=None)
    parser.add_argument("--end-date",          dest="end_date",               type=parse_iso_date, default=None)
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    args = build_args()
    if args.start_date and args.end_date and args.start_date > args.end_date:
        raise SystemExit("--start-date must be earlier than or equal to --end-date")

    global DATA_SHEET
    DATA_SHEET = args.sheet_name

    print("Reading Excel files…")
    all_tokens       = prepare_all_tokens(Path(args.all_tokens_path))
    completed_tokens = prepare_completed_tokens(Path(args.completed_tokens_path))

    if args.start_date or args.end_date:
        all_tokens       = apply_date_filter(all_tokens,       args.start_date, args.end_date)
        completed_tokens = apply_date_filter(completed_tokens, args.start_date, args.end_date)

    output_dir = Path(args.output_dir)
    print(f"\nWriting JSON to {output_dir.relative_to(BASE_DIR)}/")

    outputs = {
        "scorecards.json":             build_scorecards(all_tokens),
        "modality_status.json":        build_modality_status(all_tokens),
        "tat_vs_target.json":          build_tat_vs_target(completed_tokens),
        "tat_distribution.json":       build_tat_distribution(completed_tokens),
        "daily_trends.json":           build_daily_trends(completed_tokens),
        "daily_status_summary.json":   build_daily_status_summary(all_tokens),
        "daily_process_breakdown.json": build_daily_process_breakdown(completed_tokens),
    }

    for file_name, payload in outputs.items():
        write_json(output_dir / file_name, payload)

    print("\nDone. All JSON files updated.")


if __name__ == "__main__":
    main()